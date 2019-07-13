import sqlite3
import openpyxl
from units import unit
import units.predefined
from stravalib import unithelper as uh

import xlwings as xw
import requests
from bs4 import BeautifulSoup
import json
import brotli

def format_data(athlete_info, race_data):
    #print (race_data)
    race_distances = { "5k" : uh.meter(5000), "8k" : uh.meter(8000), "10k" : uh.meter(10000), "12k" : uh.meter(12000), "15k" : uh.meter(15000), "1/2 Mar" : uh.meter(21097.5), "Marathon" : uh.meter(42195)}

    formatted_race_data = [] 

    for race in race_data:
        for distance in race_distances:
            if (abs(race_distances[distance] - race.distance) < 0.1 * race.distance):   
                race.distance = race_distances[distance]
                # Zfill adds leading 0 if string is less than 8 characters (needed formatting for post request)
                formatted_race_data.append((race.distance, str(race.moving_time).zfill(8)))
                #Need to add way to deal with custom length race later
    return formatted_race_data
    #do more here
                

def calculate_vdot(race_data):
    
    headers = {
        'cookie': '__cfduid=d0b9f95405c5e2d2099437f23da0da88f1561088759; __utmc=140335664; _ga=GA1.2.1431627166.1561088760; __utmz=140335664.1561089351.2.2.utmcsr=google|utmccn=(organic)|utmcmd=organic|utmctr=(not%20provided); _gid=GA1.2.1220013293.1562985965; _gat=1; __utma=140335664.1431627166.1561088760.1561089351.1562985974.3; __utmt=1; __utmb=140335664.3.9.1562985987153',
        'origin': 'https://runsmartproject.com',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'en-US,en;q=0.9,de;q=0.8',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.90 Safari/537.36',
        'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'accept': 'application/json, text/javascript, */*; q=0.01',
        'referer': 'https://runsmartproject.com/calculator/',
        'authority': 'runsmartproject.com',
        'x-requested-with': 'XMLHttpRequest',
    }
    count = 0 
    for race in race_data:
        if count == 0:
            data = {
              'distance': race[0],
              'unit': 'm',
              'time': race[1],
              'pace': 'empty',
              'punit': 'mi',
              'temp': '',
              'tunit': 'F',
              'alt': '',
              'aunit': 'ft',
              'advtype': 'temperature',
              'predict': 'true'
            }

            #r = requests.get('https://runsmartproject.com/calculator/')
            response = requests.post('https://runsmartproject.com/vdot/app/api/find_paces', headers=headers, data=data)
            
            print (brotli.decompress(response.content))
            #r = requests.get('https://runsmartproject.com/calculator/')

            soup = BeautifulSoup(response.text, "html.parser")
            links = soup.find_all("a")
            tag = soup.find("div", id="vdot")
        count += 1


def create_database(data): 
    sql_create_accounts_table= """ CREATE TABLE IF NOT EXISTS race_data(
                                         id integer PRIMARY KEY,
                                         distance decimal,
                                         time decimal,
                                     ); """
    '''
    srcfile = openpyxl.load_workbook('test.xlsm',read_only=False, keep_vba= True)#to open the excel sheet and if it has macros

    sheetname = srcfile.get_sheet_by_name('Daniels Tables')#get sheetname from the file

    print ("Hi " + str(athlete_info.firstname) + " thanks for choosing our service!")


    if (athlete_info.sex == "M"):
        print ("Adding height into correct place")
        sheetname['C2'] = "72" # My height hardcoded (could have user enter it)
    elif (athlete_info.sex == "F"):
        sheetname['C2'] = "64" #Average Female Height
    else:
        sheetname['C2'] = "69" # Assume is a man if no sex listed

    if (athlete_info.dateofbirth != None):
        sheetname['E2'] = athlete_info.dateofbirth # Birthdate
    else:
        print ("Prompt user for birthdate")

    if (athlete_info.weight != None): 
        sheetname['C3'] = athlete_info.weight # Weight
    else:
        print ("Prompt user for weight")

    if (athlete_info.max_heartrate != None):
        sheetname['E3'] = athlete_info.max_heartrate# HRmax
    else:
        print ("Prompt user for heart rate max")


    uh.meter = unit('m')


    race_distances = { "5k" : uh.meter(5000), "8k" : uh.meter(8000), "10k" : uh.meter(10000), "12k" : uh.meter(12000), "15k" : uh.meter(15000), "1/2 Mar" : uh.meter(21097.5), "Marathon" : uh.meter(42195)}
    
    filled_distance = False
    for race in race_distances:
        if (abs(race_distances[race] - race_data[0].distance) < 0.1 * race_distances[race]):   
            sheetname['E6'] = race #Enter distance (need to edit S5 if Custom)
            filled_distance = True
    if (not filled_distance):
        sheetname['E6'] = "Custom" 
        sheetname['S5'] = race_data[0].distance
    

    print (sheetname['J6'].value)
    sheetname['G6'] = str(race_data[0].elapsed_time) #Edits the race time

    srcfile.save('test.xlsm')#save it as a new file, the original file is untouched and here I am saving it as xlsm(m here denotes macros).


    #wb = openpyxl.load_workbook('test.xlsm' , data_only=True)
    #data_sheet = wb['Daniels Tables']




    #srcfile.close()
    
    read_vdot()
    '''

def read_vdot():
    #srcfile = openpyxl.load_workbook('test.xlsm', read_only= True)#to open the excel sheet and if it has macros

    srcfile = openpyxl.load_workbook('test.xlsm', data_only = True)#to open the excel sheet and if it has macros
    #temp_sheetname = temp_srcfile.get_sheet_by_name('Daniels Tables')#get sheetname from the file

    temp_sheetname = srcfile.get_sheet_by_name('Daniels Tables')
    print (temp_sheetname['J6'].internal_value)
    print (temp_sheetname['C2'].value)
    print ("Filler")#Read from I/J 6 to get VDOT score
    srcfile.close()
